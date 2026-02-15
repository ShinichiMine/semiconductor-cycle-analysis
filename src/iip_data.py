"""経済産業省 鉱工業指数（IIP）データ取得・加工

電子部品・デバイス工業の出荷指数・在庫指数を取得し、
半導体サイクル（出荷/在庫比率）を算出する。

データソース: 経済産業省 鉱工業指数
https://www.meti.go.jp/statistics/tyo/iip/
"""

import os
import pandas as pd
import requests
from io import BytesIO
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent / "data"

# 経産省 鉱工業指数 公開API (e-Stat)
# 電子部品・デバイス工業: 業種コード=340
# 出荷指数・在庫指数の時系列データ

# e-Stat API用の設定
ESTAT_BASE_URL = "https://api.e-stat.go.jp/rest/3.0/app/json/getStatsData"

# IIPデータCSV保存先
IIP_CSV_PATH = DATA_DIR / "iip_electronic_parts.csv"

# e-Stat 鉱工業生産・出荷・在庫指数 statsDataId
# 2020年基準: 業種別季節調整済指数【月次】（2018年〜最新）
ESTAT_SHIP_2020 = "0004015801"  # 出荷
ESTAT_INV_2020 = "0004015802"   # 在庫
ESTAT_CAT01_ELECTRONIC = "0046000"  # 電子部品・デバイス工業

# 2015年基準: 業種別季節調整済指数【月次】（2013年〜2023年3月）
ESTAT_SHIP_2015 = "0004017875"  # 出荷
ESTAT_INV_2015 = "0004017876"   # 在庫
ESTAT_CAT02_ELECTRONIC = "1046000"  # 電子部品・デバイス工業（2015年基準）


def _fetch_estat_series(
    app_id: str, stats_data_id: str, filter_key: str, filter_value: str,
    time_field: str, time_class_id: str,
) -> pd.Series:
    """e-Stat APIから1系列を取得する共通関数"""
    params = {
        "appId": app_id,
        "lang": "J",
        "statsDataId": stats_data_id,
        f"cd{filter_key.title()}": filter_value,
        "metaGetFlg": "Y",
        "cntGetFlg": "N",
    }
    resp = requests.get(ESTAT_BASE_URL, params=params, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    result = data.get("GET_STATS_DATA", {}).get("RESULT", {})
    if result.get("STATUS") != 0:
        raise ValueError(f"e-Stat API error: {result}")

    stat_data = data.get("GET_STATS_DATA", {}).get("STATISTICAL_DATA", {})

    # Build time code -> YYYYMM name mapping
    time_map = {}
    for cls in stat_data.get("CLASS_INF", {}).get("CLASS_OBJ", []):
        if cls.get("@id") == time_class_id:
            items = cls.get("CLASS", [])
            if isinstance(items, dict):
                items = [items]
            for item in items:
                time_map[item.get("@code", "")] = item.get("@name", "")

    values = stat_data.get("DATA_INF", {}).get("VALUE", [])
    records = {}
    for v in values:
        time_code = v.get(f"@{time_field}", "")
        time_name = time_map.get(time_code, "")
        val = v.get("$", "")
        if val in ("", "-", "***", "x") or not time_name:
            continue
        if len(time_name) == 6 and time_name.isdigit():
            try:
                records[pd.to_datetime(time_name, format="%Y%m")] = float(val)
            except (ValueError, TypeError):
                continue
    return pd.Series(records).sort_index()


def fetch_iip_from_estat(app_id: str | None = None) -> pd.DataFrame | None:
    """e-Stat APIから電子部品・デバイス工業のIIPデータを取得し、CSVに保存する。

    2015年基準（2013年〜2023年3月）と2020年基準（2018年〜最新）を結合し、
    重複期間の比率でリベースして長期データを構築する。

    環境変数 ESTAT_APP_ID、または引数 app_id でアプリケーションIDを指定する。
    取得成功時は data/iip_electronic_parts.csv に保存し、DataFrameを返す。

    Returns:
        DataFrame (columns: shipment_index, inventory_index, index=date) or None
    """
    app_id = app_id or os.environ.get("ESTAT_APP_ID")
    if not app_id:
        return None

    try:
        # 2020年基準（最新データ）
        ship20 = _fetch_estat_series(
            app_id, ESTAT_SHIP_2020, "cat01", ESTAT_CAT01_ELECTRONIC, "time", "time")
        inv20 = _fetch_estat_series(
            app_id, ESTAT_INV_2020, "cat01", ESTAT_CAT01_ELECTRONIC, "time", "time")

        # 2015年基準（過去データ）
        ship15 = _fetch_estat_series(
            app_id, ESTAT_SHIP_2015, "cat02", ESTAT_CAT02_ELECTRONIC, "cat01", "cat01")
        inv15 = _fetch_estat_series(
            app_id, ESTAT_INV_2015, "cat02", ESTAT_CAT02_ELECTRONIC, "cat01", "cat01")

        # 重複期間でリベース比率算出
        overlap = ship15.index.intersection(ship20.index)
        if len(overlap) < 6:
            # 重複不十分なら2020年基準のみ使用
            df = pd.DataFrame({
                "shipment_index": ship20, "inventory_index": inv20,
            }).dropna()
        else:
            ship_ratio = ship20[overlap].mean() / ship15[overlap].mean()
            inv_ratio = inv20[overlap].mean() / inv15[overlap].mean()

            ship15_rebased = ship15 * ship_ratio
            inv15_rebased = inv15 * inv_ratio

            ship_combined = pd.concat([
                ship15_rebased[ship15_rebased.index < ship20.index[0]], ship20])
            inv_combined = pd.concat([
                inv15_rebased[inv15_rebased.index < inv20.index[0]], inv20])

            df = pd.DataFrame({
                "shipment_index": ship_combined, "inventory_index": inv_combined,
            }).dropna()

        df.index.name = "date"
        df = df.sort_index()

        DATA_DIR.mkdir(parents=True, exist_ok=True)
        df.to_csv(IIP_CSV_PATH)
        return df

    except Exception:
        return None


def fetch_iip_from_meti_excel(timeout: int = 30) -> pd.DataFrame | None:
    """経産省サイトからIIP Excelファイルを自動DLし、電子部品・デバイス工業データを抽出する。

    URLパターン: https://www.meti.go.jp/statistics/tyo/iip/result/xlsx/b2020_sj{X}data{Y}.xlsx
    X=1(出荷), X=2(在庫)  Y=1(原数値), Y=2(季節調整済)

    Args:
        timeout: HTTPリクエストのタイムアウト秒数。

    Returns:
        DataFrame (columns: shipment_index, inventory_index, index=date) or None
    """
    try:
        import openpyxl
    except ImportError:
        print("  [WARN] openpyxl未インストール。pip install openpyxl で追加してください。")
        return None

    base_url = "https://www.meti.go.jp/statistics/tyo/iip/result/xlsx"

    # 季節調整済指数: X=1(出荷), X=2(在庫), Y=2(季節調整済)
    urls = {
        "shipment": f"{base_url}/b2020_sj1data2.xlsx",
        "inventory": f"{base_url}/b2020_sj2data2.xlsx",
    }

    series = {}
    for key, url in urls.items():
        try:
            print(f"  経産省Excel DL中: {key}...")
            resp = requests.get(url, timeout=timeout)
            resp.raise_for_status()
            wb = openpyxl.load_workbook(BytesIO(resp.content), data_only=True)

            # 「電子部品・デバイス工業」を含む行を探索
            df_found = None
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
                    row_values = [cell.value for cell in row]
                    # 電子部品・デバイス工業を含む行を探す
                    if any("電子部品" in str(v) for v in row_values if v):
                        # この行のデータを抽出（ヘッダー行から日付を取得する必要あり）
                        # ヘッダー行を探す（年月の数値列が並ぶ行）
                        header_row = None
                        for hr in ws.iter_rows(min_row=1, max_row=row[0].row, values_only=False):
                            hr_vals = [cell.value for cell in hr]
                            # 年月を示す数値が多い行をヘッダーとする
                            date_count = sum(1 for v in hr_vals
                                           if isinstance(v, (int, float)) and 200000 < v < 210000)
                            if date_count == 0:
                                # 「年」「月」のテキストがある行
                                date_count = sum(1 for v in hr_vals
                                               if isinstance(v, str) and ("年" in v or "月" in v))
                            if date_count >= 3:
                                header_row = hr
                                break

                        if header_row is None:
                            continue

                        records = {}
                        for hi, cell in enumerate(header_row):
                            hval = cell.value
                            data_val = row[hi].value if hi < len(row) else None
                            if data_val is None or not isinstance(data_val, (int, float)):
                                continue
                            # 日付パース
                            date = None
                            if isinstance(hval, (int, float)) and 200000 < hval < 210000:
                                try:
                                    date = pd.to_datetime(str(int(hval)), format="%Y%m")
                                except ValueError:
                                    continue
                            if date:
                                records[date] = float(data_val)

                        if records:
                            series[key] = pd.Series(records).sort_index()
                            break
                if key in series:
                    break

        except Exception as e:
            print(f"  [WARN] 経産省Excel取得失敗 ({key}): {e}")
            return None

    if "shipment" not in series or "inventory" not in series:
        print("  [WARN] 経産省Excelから電子部品・デバイス工業データを抽出できませんでした。")
        return None

    df = pd.DataFrame({
        "shipment_index": series["shipment"],
        "inventory_index": series["inventory"],
    }).dropna()
    df.index.name = "date"
    df = df.sort_index()

    # キャッシュ保存
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(IIP_CSV_PATH)
    print(f"  経産省Excelからデータ取得成功: {df.index.min()} 〜 {df.index.max()}")
    return df


def parse_meti_excel(source) -> pd.DataFrame | None:
    """経産省サイトからDLした生Excel（b2020_gsm1j.xlsx等）をパースする。

    Args:
        source: ファイルパス（str/Path）またはバイナリデータ（bytes/BytesIO）

    Returns:
        DataFrame (columns: shipment_index, inventory_index, index=date) or None
    """
    try:
        import openpyxl
    except ImportError:
        print("  [WARN] openpyxl未インストール。pip install openpyxl で追加してください。")
        return None

    try:
        if isinstance(source, (str, Path)):
            wb = openpyxl.load_workbook(source, data_only=True)
        else:
            wb = openpyxl.load_workbook(BytesIO(source) if isinstance(source, bytes) else source, data_only=True)

        # シート名の判定（月次: 出荷/在庫、四半期: 出荷計/在庫計（期末））
        sheet_map = {}
        for name in wb.sheetnames:
            if "出荷計" in name:
                sheet_map.setdefault("shipment", name)
            elif "出荷" in name:
                sheet_map["shipment"] = name  # 月次優先
            if "在庫計" in name:
                sheet_map.setdefault("inventory", name)
            elif "在庫" in name:
                sheet_map["inventory"] = name  # 月次優先

        if "shipment" not in sheet_map or "inventory" not in sheet_map:
            print(f"  [WARN] 出荷/在庫シートが見つかりません。シート一覧: {wb.sheetnames}")
            return None

        series = {}
        for key in ("shipment", "inventory"):
            ws = wb[sheet_map[key]]

            # R3（3行目）をヘッダーとして読み取る
            header_row = [cell.value for cell in ws[3]]

            # 日付列のインデックスを特定（int YYYYMM or "p YYYYMM"文字列）
            date_cols = {}  # col_index -> datetime
            for ci, hval in enumerate(header_row):
                dt = None
                if isinstance(hval, (int, float)) and 200000 < hval < 210000:
                    try:
                        dt = pd.to_datetime(str(int(hval)), format="%Y%m")
                    except ValueError:
                        pass
                elif isinstance(hval, str):
                    # "p 202512" のような速報値ヘッダー
                    cleaned = hval.strip().lstrip("p").strip()
                    if cleaned.isdigit() and len(cleaned) == 6:
                        try:
                            dt = pd.to_datetime(cleaned, format="%Y%m")
                        except ValueError:
                            pass
                if dt is not None:
                    date_cols[ci] = dt

            if not date_cols:
                print(f"  [WARN] シート'{sheet_map[key]}'で日付列を検出できませんでした。")
                return None

            # 「電子部品・デバイス工業」行を探す
            target_row = None
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
                row_values = [cell.value for cell in row]
                for v in row_values:
                    if v and "電子部品" in str(v):
                        target_row = row_values
                        break
                if target_row is not None:
                    break

            if target_row is None:
                print(f"  [WARN] シート'{sheet_map[key]}'で「電子部品・デバイス工業」行が見つかりません。")
                return None

            # データ抽出
            records = {}
            for ci, dt in date_cols.items():
                if ci < len(target_row):
                    val = target_row[ci]
                    if isinstance(val, (int, float)):
                        records[dt] = float(val)

            if not records:
                print(f"  [WARN] シート'{sheet_map[key]}'からデータを抽出できませんでした。")
                return None

            series[key] = pd.Series(records).sort_index()

        # DataFrame構築
        df_new = pd.DataFrame({
            "shipment_index": series["shipment"],
            "inventory_index": series["inventory"],
        }).dropna()
        df_new.index.name = "date"
        df_new = df_new.sort_index()

        if df_new.empty:
            print("  [WARN] 経産省Excelから有効なデータが抽出できませんでした。")
            return None

        # 既存キャッシュがあれば結合（2013-2017年分を保持し、2018年以降は新データで上書き）
        if IIP_CSV_PATH.exists():
            try:
                cached = load_iip_from_csv(IIP_CSV_PATH)
                new_start = df_new.index.min()
                old_part = cached[cached.index < new_start]
                if not old_part.empty:
                    df_new = pd.concat([old_part, df_new]).sort_index()
                    print(f"  既存キャッシュ({old_part.index.min().strftime('%Y-%m')}〜{old_part.index.max().strftime('%Y-%m')})と結合しました。")
            except Exception:
                pass  # 結合失敗しても新データだけで続行

        # キャッシュ保存
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        df_new.to_csv(IIP_CSV_PATH)
        print(f"  経産省Excel解析成功: {df_new.index.min().strftime('%Y-%m')} 〜 {df_new.index.max().strftime('%Y-%m')}（{len(df_new)}件）")
        return df_new

    except Exception as e:
        print(f"  [ERROR] 経産省Excelパース失敗: {e}")
        return None


def load_iip_from_upload(contents: str, filename: str) -> pd.DataFrame | None:
    """ダッシュボードのアップロード機能からIIPデータを読み込む。

    Args:
        contents: base64エンコードされたファイルコンテンツ（dcc.Upload形式）
        filename: アップロードされたファイル名

    Returns:
        DataFrame (columns: shipment_index, inventory_index, index=date) or None
    """
    import base64

    try:
        content_type, content_string = contents.split(",")
        decoded = base64.b64decode(content_string)

        if filename.endswith((".xlsx", ".xls")):
            # 経産省Excel生フォーマットの自動判定
            try:
                import openpyxl
                wb = openpyxl.load_workbook(BytesIO(decoded), data_only=True)
                if any("出荷" in name for name in wb.sheetnames):
                    # 経産省フォーマット → parse_meti_excel()へ
                    print(f"  経産省Excelフォーマットを検出: {filename}")
                    wb.close()
                    df = parse_meti_excel(decoded)
                    if df is not None:
                        return df
                    return None
                wb.close()
            except Exception:
                pass
            # 通常のExcel（date/shipment_index/inventory_index列）
            df = pd.read_excel(BytesIO(decoded), parse_dates=["date"], index_col="date")
        elif filename.endswith(".csv"):
            df = pd.read_csv(BytesIO(decoded), parse_dates=["date"], index_col="date")
        else:
            print(f"  [WARN] 未対応ファイル形式: {filename}")
            return None

        # 必須列の確認
        required = {"shipment_index", "inventory_index"}
        if not required.issubset(df.columns):
            print(f"  [WARN] 必須列不足。必要: {required}、実際: {set(df.columns)}")
            return None

        df = df[["shipment_index", "inventory_index"]].dropna()
        df.index.name = "date"

        # キャッシュ保存
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        df.to_csv(IIP_CSV_PATH)
        return df

    except Exception as e:
        print(f"  [ERROR] アップロードファイル読み込み失敗: {e}")
        return None


def load_or_fetch_iip(app_id: str | None = None) -> pd.DataFrame:
    """IIPデータを取得する（優先度: ローカルCSV → e-Stat API → 経産省Excel → サンプルデータ）

    1. data/iip_electronic_parts.csv が存在する場合、最終データ日付を確認。
       - 最終データ日付が30日以上前 AND 取得手段あり: API再取得 → Excel自動DL を試みる。
         成功 → 新CSVで上書き＋返す。失敗 → 既存CSVを返す（フォールバック）。
       - 最終データ日付が30日以内 → 既存CSVをそのまま返す。
    2. CSVが存在しなければ e-Stat API → 経産省Excel自動DL → サンプルデータの順で試行。

    Returns:
        DataFrame (columns: shipment_index, inventory_index, index=date)
    """
    app_id = app_id or os.environ.get("ESTAT_APP_ID")

    if IIP_CSV_PATH.exists():
        cached = load_iip_from_csv(IIP_CSV_PATH)
        last_date = cached.index.max()
        days_old = (pd.Timestamp.now() - last_date).days
        if days_old > 30:
            print(f"IIPキャッシュが{days_old}日前のデータ。再取得を試みます...")
            if app_id:
                fetched = fetch_iip_from_estat(app_id=app_id)
                if fetched is not None:
                    print(f"IIPデータ更新成功（e-Stat）: {fetched.index.min()} 〜 {fetched.index.max()}")
                    return fetched
                print("e-Stat API取得失敗。経産省Excelを試みます...")
            fetched = fetch_iip_from_meti_excel(timeout=10)
            if fetched is not None:
                return fetched
            print("全経路失敗。既存キャッシュを使用します。")
        return cached

    # CSVなし: 順次試行
    if app_id:
        fetched = fetch_iip_from_estat(app_id=app_id)
        if fetched is not None:
            return fetched

    fetched = fetch_iip_from_meti_excel()
    if fetched is not None:
        return fetched

    return create_sample_iip_data()


def load_iip_from_csv(filepath: str | Path) -> pd.DataFrame:
    """ローカルCSVからIIPデータを読み込む

    経産省サイトから手動DLしたCSVを想定。
    列: date, shipment_index, inventory_index
    """
    df = pd.read_csv(filepath, parse_dates=["date"], index_col="date")
    return df


def create_sample_iip_data() -> pd.DataFrame:
    """分析用サンプルデータ生成（実データ取得までの代替）

    経産省IIP「電子部品・デバイス工業」の出荷・在庫指数の
    実績値に基づく近似データ。半導体サイクルの特徴を再現:
    - 2010-2011: 回復期
    - 2012-2013: 調整期
    - 2014-2016: 緩やかな回復
    - 2017-2018: スーパーサイクル
    - 2019: 調整
    - 2020-2021: コロナ需要
    - 2022-2023: 在庫調整
    - 2024-2025: AI需要回復
    """
    import numpy as np

    dates = pd.date_range("2010-01-01", "2025-12-31", freq="ME")
    n = len(dates)
    t = np.arange(n)

    # 基調トレンド（緩やかな上昇）
    trend = 95 + t * 0.08

    # 半導体サイクル（約3-4年周期）
    cycle = 12 * np.sin(2 * np.pi * t / 42)

    # 短期変動
    np.random.seed(42)
    noise = np.random.normal(0, 2, n)

    # イベント効果
    event = np.zeros(n)
    # 2017-2018 スーパーサイクル
    event[84:108] += 8
    # 2020 コロナショック → 回復
    event[120:126] -= 10
    event[126:144] += 12
    # 2022-2023 在庫調整
    event[144:162] -= 6
    # 2024-2025 AI需要
    event[168:] += 10

    shipment = trend + cycle + event + noise
    shipment = np.clip(shipment, 70, 140)

    # 在庫は出荷の逆位相（3-6ヶ月遅れ）
    inventory_base = 100 + t * 0.05
    inventory_cycle = -8 * np.sin(2 * np.pi * (t - 5) / 42)
    inventory_event = np.zeros(n)
    inventory_event[126:150] += 15  # コロナ後の在庫積み上がり
    inventory_event[150:168] += 8
    inventory_event[168:] -= 5  # AI需要で在庫消化

    inventory = inventory_base + inventory_cycle + inventory_event + noise * 0.8
    inventory = np.clip(inventory, 75, 135)

    df = pd.DataFrame(
        {"shipment_index": shipment, "inventory_index": inventory},
        index=dates,
    )
    df.index.name = "date"
    return df


def calc_cycle_indicator(df: pd.DataFrame) -> pd.DataFrame:
    """半導体サイクル指標を算出

    Returns:
        DataFrame with columns:
        - shipment_index: 出荷指数
        - inventory_index: 在庫指数
        - si_ratio: 出荷/在庫比率（>1で好況局面）
        - si_ratio_ma3: 出荷/在庫比率の3ヶ月移動平均
        - shipment_yoy: 出荷指数の前年同月比(%)
        - inventory_yoy: 在庫指数の前年同月比(%)
        - cycle_phase: サイクル局面判定
    """
    result = df.copy()

    # 出荷/在庫比率
    result["si_ratio"] = result["shipment_index"] / result["inventory_index"]
    result["si_ratio_ma3"] = result["si_ratio"].rolling(3).mean()

    # 前年同月比
    result["shipment_yoy"] = result["shipment_index"].pct_change(12) * 100
    result["inventory_yoy"] = result["inventory_index"].pct_change(12) * 100

    # サイクル局面判定
    result["cycle_phase"] = _classify_phase(result)

    return result


def _classify_phase(df: pd.DataFrame) -> pd.Series:
    """出荷/在庫比率と変化方向からサイクル局面を4分類

    1. 回復期 (Recovery):  SI比率上昇 & SI比率<1
    2. 好況期 (Expansion): SI比率上昇 & SI比率>=1
    3. 後退期 (Slowdown):  SI比率下降 & SI比率>=1
    4. 不況期 (Recession):  SI比率下降 & SI比率<1
    """
    si = df["si_ratio_ma3"]
    si_diff = si.diff()

    conditions = []
    for i in range(len(df)):
        if pd.isna(si.iloc[i]) or pd.isna(si_diff.iloc[i]):
            conditions.append("不明")
        elif si_diff.iloc[i] >= 0 and si.iloc[i] < 1:
            conditions.append("回復期")
        elif si_diff.iloc[i] >= 0 and si.iloc[i] >= 1:
            conditions.append("好況期")
        elif si_diff.iloc[i] < 0 and si.iloc[i] >= 1:
            conditions.append("後退期")
        else:
            conditions.append("不況期")

    return pd.Series(conditions, index=df.index, name="cycle_phase")
